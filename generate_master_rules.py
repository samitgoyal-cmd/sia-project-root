#!/usr/bin/env python3
"""
SEM Master Rules Sheet Generator
Consolidates rules from Architecture v4.2 + SEM ORD-01 through ORD-07
Outputs a 22-column Excel spreadsheet with comprehensive rule consolidation
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
from datetime import datetime

# Define the workbook and sheet
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Create Summary Sheet
summary_sheet = wb.create_sheet("Summary", 0)
rules_sheet = wb.create_sheet("Master Rules", 1)

# Define columns for Master Rules Sheet
columns = [
    "Master_Rule_ID",          # MR-001, MR-002, etc.
    "Rule_Type",                # MUST, NEVER, RULE, POLICY, DEF, NOTE, ESCALATION
    "Priority",                 # CRITICAL, HIGH, MEDIUM, LOW
    "Layer",                    # Foundation, Identity-Roles, Actions-Activities, Validation-Integrity, Gamification, Outputs, Exceptions-Escalations
    "Source_Document",          # Architecture v4.2, SEM-ORD-01, etc.
    "Source_Rule_ID",           # Original rule ID(s) with line references
    "Rule_Title",               # Brief title/heading
    "Rule_Description",         # Full rule description
    "Applies_To",               # Roles, systems, layers affected
    "Conflict_Flag",            # Yes/No - indicates conflicts with other rules
    "Gap_Flag",                 # Yes/No - indicates missing implementation details
    "Edge_Case_Flag",           # Yes/No - indicates edge case or boundary condition
    "Depends_On",               # Rule IDs this depends on
    "Impacts",                  # Rule IDs this impacts
    "Duplicate_Group_ID",       # DG-001, DG-002 for deduplicated rules
    "Dedup_Status",             # Master, Duplicate, Variant
    "Implementation_Notes",     # Technical notes for implementation
    "Cross_Reference",          # Links to other documents
    "Requires_Config",          # Config items this rule depends on
    "Scope_Limitation",         # Version constraints (v1, future, etc.)
    "Authority_Level",          # Frozen, Active, Draft, Candidate
    "Last_Modified",            # Date of last modification/freeze
]

# Set up header row
for col_idx, col_name in enumerate(columns, 1):
    cell = rules_sheet.cell(row=1, column=col_idx)
    cell.value = col_name
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Set column widths
column_widths = [12, 12, 10, 18, 16, 20, 20, 35, 20, 12, 10, 12, 15, 15, 15, 15, 20, 20, 15, 15, 15, 12]
for col_idx, width in enumerate(column_widths, 1):
    rules_sheet.column_dimensions[get_column_letter(col_idx)].width = width

# Define color mapping for rule types
color_map = {
    "MUST": "FF0000",           # Red
    "NEVER": "8B0000",          # Dark Red
    "RULE": "0070C0",           # Blue
    "POLICY": "00B050",         # Green
    "DEF": "D9D9D9",            # Light Grey
    "NOTE": "E8E8E8",           # Very Light Grey
    "ESCALATION": "FF9900",     # Orange
}

# Layer definitions for 7-layer organization
layer_map = {
    "Foundation": "01_Foundation",
    "Identity-Roles": "02_Identity_Roles",
    "Actions-Activities": "03_Actions_Activities",
    "Validation-Integrity": "04_Validation_Integrity",
    "Gamification": "05_Gamification",
    "Outputs": "06_Outputs",
    "Exceptions-Escalations": "07_Exceptions_Escalations",
}

# Master rules data - consolidated from all sources
master_rules = [
    # Architecture v4.2 Core Rules (Rules 16-23 are v4.2 additions per authority)
    {
        "Master_Rule_ID": "MR-001",
        "Rule_Type": "MUST",
        "Priority": "CRITICAL",
        "Layer": "Foundation",
        "Source_Document": "Architecture v4.2",
        "Source_Rule_ID": "ARCH-16",
        "Rule_Title": "SEM is the engagement layer of SIA platform",
        "Rule_Description": "The SEM (SIA Engagement Module) is the canonical engagement, progression, and community layer for SIA. All user participation, role progression, and community activity is managed through SEM.",
        "Applies_To": "All users, All roles, All modules",
        "Conflict_Flag": "No",
        "Gap_Flag": "No",
        "Edge_Case_Flag": "No",
        "Depends_On": "",
        "Impacts": "All ORD documents",
        "Duplicate_Group_ID": "None",
        "Dedup_Status": "Master",
        "Implementation_Notes": "Highest authority governance document. Non-negotiable.",
        "Cross_Reference": "SIA_Overall_Platform_Architecture_v4.2.md",
        "Requires_Config": "None",
        "Scope_Limitation": "All versions",
        "Authority_Level": "Frozen",
        "Last_Modified": "2026-04-12",
    },
    {
        "Master_Rule_ID": "MR-002",
        "Rule_Type": "MUST",
        "Priority": "CRITICAL",
        "Layer": "Foundation",
        "Source_Document": "Architecture v4.2",
        "Source_Rule_ID": "ARCH-17",
        "Rule_Title": "No user creation outside SEM registration flow",
        "Rule_Description": "Users can only enter the SIA ecosystem through the SEM registration and onboarding flow. No backdoor user creation or bulk import bypassing SEM.",
        "Applies_To": "All users, Registration, Onboarding",
        "Conflict_Flag": "No",
        "Gap_Flag": "No",
        "Edge_Case_Flag": "No",
        "Depends_On": "MR-001",
        "Impacts": "SEM-ORD-01, SEM-ORD-07",
        "Duplicate_Group_ID": "None",
        "Dedup_Status": "Master",
        "Implementation_Notes": "Ensures all users go through standardized entry process.",
        "Cross_Reference": "SEM-ORD-01.1, SEM-ORD-07.19",
        "Requires_Config": "None",
        "Scope_Limitation": "All versions",
        "Authority_Level": "Frozen",
        "Last_Modified": "2026-04-12",
    },
    {
        "Master_Rule_ID": "MR-003",
        "Rule_Type": "NEVER",
        "Priority": "CRITICAL",
        "Layer": "Validation-Integrity",
        "Source_Document": "Architecture v4.2",
        "Source_Rule_ID": "ARCH-18",
        "Rule_Title": "Points NEVER awarded for passive presence",
        "Rule_Description": "Gamification points are NEVER awarded for passive attendance, viewing, or membership alone. Points are awarded only for active contribution, creation, or validated participation.",
        "Applies_To": "Gamification, Points system, All roles",
        "Conflict_Flag": "No",
        "Gap_Flag": "No",
        "Edge_Case_Flag": "No",
        "Depends_On": "MR-001",
        "Impacts": "SEM-ORD-04, SEM-ORD-05",
        "Duplicate_Group_ID": "DG-001",
        "Dedup_Status": "Master",
        "Implementation_Notes": "Core anti-gaming principle. Repeated across multiple ORDs.",
        "Cross_Reference": "SEM-ORD-04.10.003, SEM-ORD-05.3.004",
        "Requires_Config": "Points calculation rules",
        "Scope_Limitation": "v1+",
        "Authority_Level": "Frozen",
        "Last_Modified": "2026-04-12",
    },
    {
        "Master_Rule_ID": "MR-004",
        "Rule_Type": "RULE",
        "Priority": "CRITICAL",
        "Layer": "Identity-Roles",
        "Source_Document": "SEM-ORD-01",
        "Source_Rule_ID": "SEM-ORD-01.1.002",
        "Rule_Title": "SEM governs all user role definitions",
        "Rule_Description": "The SEM module defines and governs all user roles, role transitions, role privileges, and role hierarchy within the SIA platform.",
        "Applies_To": "All roles, Role progression, Permission model",
        "Conflict_Flag": "No",
        "Gap_Flag": "No",
        "Edge_Case_Flag": "No",
        "Depends_On": "MR-001",
        "Impacts": "SEM-ORD-01, SEM-ORD-02, SEM-ORD-07",
        "Duplicate_Group_ID": "None",
        "Dedup_Status": "Master",
        "Implementation_Notes": "Authoritative source for role definitions.",
        "Cross_Reference": "SEM-ORD-01.1.002",
        "Requires_Config": "Role definitions table",
        "Scope_Limitation": "v1+",
        "Authority_Level": "Frozen",
        "Last_Modified": "2026-04-13",
    },
    {
        "Master_Rule_ID": "MR-005",
        "Rule_Type": "NEVER",
        "Priority": "HIGH",
        "Layer": "Validation-Integrity",
        "Source_Document": "SEM-ORD-01",
        "Source_Rule_ID": "SEM-ORD-01.4.003",
        "Rule_Title": "SIA IDs are permanent and immutable",
        "Rule_Description": "A SIA ID, once assigned to a user, can never be changed, deleted, or transferred. SIA IDs are permanent identifiers for life of the account.",
        "Applies_To": "All users, Identity system",
        "Conflict_Flag": "No",
        "Gap_Flag": "No",
        "Edge_Case_Flag": "No",
        "Depends_On": "MR-004",
        "Impacts": "SEM-ORD-01.4",
        "Duplicate_Group_ID": "None",
        "Dedup_Status": "Master",
        "Implementation_Notes": "Database constraint required at schema level.",
        "Cross_Reference": "SEM-ORD-01.4.003",
        "Requires_Config": "Database schema",
        "Scope_Limitation": "v1+",
        "Authority_Level": "Frozen",
        "Last_Modified": "2026-04-13",
    },
]

# Add more rules programmatically from each ORD document
# This is a sample - in production, extract all ~600+ rules

# Helper function to add rules
def add_rule(mr_id, rule_type, priority, layer, source_doc, source_id, title, desc, applies_to,
             depends="", impacts="", dup_group="None", dup_status="Master"):
    master_rules.append({
        "Master_Rule_ID": mr_id,
        "Rule_Type": rule_type,
        "Priority": priority,
        "Layer": layer,
        "Source_Document": source_doc,
        "Source_Rule_ID": source_id,
        "Rule_Title": title,
        "Rule_Description": desc,
        "Applies_To": applies_to,
        "Conflict_Flag": "No",
        "Gap_Flag": "No",
        "Edge_Case_Flag": "No",
        "Depends_On": depends,
        "Impacts": impacts,
        "Duplicate_Group_ID": dup_group,
        "Dedup_Status": dup_status,
        "Implementation_Notes": "",
        "Cross_Reference": source_id,
        "Requires_Config": "",
        "Scope_Limitation": "v1+",
        "Authority_Level": "Frozen",
        "Last_Modified": "2026-04-13",
    })

# Add sample rules from ORD-01 (Role Framework)
add_rule("MR-006", "DEF", "HIGH", "Identity-Roles", "SEM-ORD-01", "SEM-ORD-01.2.001",
    "BIT Club Member definition",
    "A BIT Club Member is a registered SIA user who has joined the BIT Club participation tier.",
    "All users who join BIT Club")

add_rule("MR-007", "RULE", "HIGH", "Identity-Roles", "SEM-ORD-01", "SEM-ORD-01.3.011",
    "BIT Club members can participate in forum",
    "BIT Club members gain access to forum participation and can post in designated zones.",
    "BIT Club members, Forum")

add_rule("MR-008", "DEF", "HIGH", "Identity-Roles", "SEM-ORD-01", "SEM-ORD-01.7.001",
    "Campus Rep role definition",
    "A Campus Rep is a user responsible for representing SIA at a specific educational campus and conducting campus visits.",
    "Campus representatives, Field operations")

add_rule("MR-009", "NEVER", "CRITICAL", "Identity-Roles", "SEM-ORD-01", "SEM-ORD-01.7.020",
    "Campus Rep cannot approve own events",
    "A Campus Rep can never approve or validate their own events or contributions.",
    "Campus Rep, Events, Validation")

add_rule("MR-010", "RULE", "HIGH", "Actions-Activities", "SEM-ORD-02", "SEM-ORD-02.3.001",
    "Campus visits are structured activation events",
    "Campus visits are deliberate, planned activation events that follow a standardized flow of preparation, execution, and follow-up.",
    "Campus visits, Field operations")

# Add sample rules from ORD-03 (Forum)
add_rule("MR-011", "NEVER", "CRITICAL", "Validation-Integrity", "SEM-ORD-03", "SEM-ORD-03.3.003",
    "Forum content cannot be sold or commercialized",
    "Forum discussions and user-generated content are never monetized or sold to third parties. Content remains the community property.",
    "Forum, Content protection")

add_rule("MR-012", "RULE", "MEDIUM", "Outputs", "SEM-ORD-03", "SEM-ORD-03.12.002",
    "Forum contributions generate scorable activity",
    "Substantive forum posts and accepted answers generate scoring points that flow to the gamification engine.",
    "Forum, Gamification, Scoring")

# Add sample rules from ORD-04 (Events)
add_rule("MR-013", "NEVER", "CRITICAL", "Validation-Integrity", "SEM-ORD-04", "SEM-ORD-04.3.003",
    "Events NEVER reward attendance inflation",
    "An event with inflated attendee counts without substantive engagement, no report, and no contributor validation is not a successful SIA event.",
    "Events, Validation")

add_rule("MR-014", "RULE", "HIGH", "Actions-Activities", "SEM-ORD-04", "SEM-ORD-04.5.001",
    "Standard event lifecycle",
    "All events follow: Draft → Review → Approval → Execution → Report → Validation → Closure",
    "Events, Lifecycle")

add_rule("MR-015", "NEVER", "HIGH", "Actions-Activities", "SEM-ORD-04", "SEM-ORD-04.11.005",
    "Events without reports generate no credit",
    "An event where no Event Report is submitted NEVER generates event credit, regardless of whether the event occurred.",
    "Events, Reporting, Gamification")

# Add sample rules from ORD-05 (Gamification)
add_rule("MR-016", "POLICY", "HIGH", "Gamification", "SEM-ORD-05", "SEM-ORD-05.3.001",
    "Gamification should motivate contribution",
    "The gamification system is designed to motivate genuine contribution, knowledge sharing, and community growth.",
    "Gamification, Motivation")

add_rule("MR-017", "RULE", "MEDIUM", "Gamification", "SEM-ORD-05", "SEM-ORD-05.5.001",
    "Points are tracked per activity category",
    "Points are accumulated separately by activity category (Forum, Events, Tasks, Books, etc.) and rolled up to total score.",
    "Points, Gamification, Scoring")

add_rule("MR-018", "ESCALATION", "MEDIUM", "Exceptions-Escalations", "SEM-ORD-05", "SEM-ORD-05.14.006",
    "Suspicious score patterns escalate to Founder",
    "Any user demonstrating suspicious patterns (rapid score gains, reciprocal gaming, etc.) escalates for Founder review.",
    "Gamification, Anti-gaming, Escalation")

# Add sample rules from ORD-06 (Certification)
add_rule("MR-019", "DEF", "HIGH", "Outputs", "SEM-ORD-06", "SEM-ORD-06.2.001",
    "Certificate definition",
    "A Certificate is a digital or printable credential issued to a user upon completion of a defined role, activity, or milestone.",
    "Certification, Credentials")

add_rule("MR-020", "RULE", "HIGH", "Outputs", "SEM-ORD-06", "SEM-ORD-06.5.001",
    "Certificates triggered on role/activity completion",
    "Certificates are automatically triggered and queued upon completion of defined roles, activities, or milestones.",
    "Certification, Automation")

# Add sample rules from ORD-07 (Forms, Dashboards, Onboarding)
add_rule("MR-021", "RULE", "MEDIUM", "Outputs", "SEM-ORD-07", "SEM-ORD-07.7.001",
    "Form approval matrix governs submission processing",
    "Each form type follows an approval matrix specifying who can auto-approve, review, or escalate submissions.",
    "Forms, Approval, Workflow")

add_rule("MR-022", "RULE", "MEDIUM", "Outputs", "SEM-ORD-07", "SEM-ORD-07.9.002",
    "Dashboards show user's current position and next steps",
    "Role-based dashboards clearly show users their current position, earned points/badges, and the next available actions/milestones.",
    "Dashboards, UI, User experience")

# Populate rules sheet with master rules
for row_idx, rule in enumerate(master_rules, 2):
    for col_idx, col_name in enumerate(columns, 1):
        cell = rules_sheet.cell(row=row_idx, column=col_idx)
        cell.value = rule.get(col_name, "")

        # Apply color based on rule type
        rule_type = rule.get("Rule_Type", "")
        if rule_type in color_map:
            cell.fill = PatternFill(start_color=color_map[rule_type], end_color=color_map[rule_type], fill_type="solid")
            if rule_type in ["MUST", "NEVER"]:
                cell.font = Font(color="FFFFFF")

        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

# Freeze header row
rules_sheet.freeze_panes = "A2"

# Create Summary Sheet
summary_sheet.column_dimensions["A"].width = 30
summary_sheet.column_dimensions["B"].width = 20

title_cell = summary_sheet.cell(row=1, column=1)
title_cell.value = "SEM Master Rules Sheet - Summary"
title_cell.font = Font(bold=True, size=14)

# Add summary statistics
summary_sheet.cell(row=3, column=1).value = "Summary Statistics"
summary_sheet.cell(row=3, column=1).font = Font(bold=True, size=12)

summary_sheet.cell(row=4, column=1).value = "Total Rules:"
summary_sheet.cell(row=4, column=2).value = len(master_rules)

rule_types = {}
for rule in master_rules:
    rt = rule.get("Rule_Type", "Unknown")
    rule_types[rt] = rule_types.get(rt, 0) + 1

summary_sheet.cell(row=6, column=1).value = "Rules by Type"
summary_sheet.cell(row=6, column=1).font = Font(bold=True)

row_offset = 7
for rule_type, count in sorted(rule_types.items()):
    summary_sheet.cell(row=row_offset, column=1).value = rule_type
    summary_sheet.cell(row=row_offset, column=2).value = count
    if rule_type in color_map:
        summary_sheet.cell(row=row_offset, column=1).fill = PatternFill(
            start_color=color_map[rule_type], end_color=color_map[rule_type], fill_type="solid"
        )
    row_offset += 1

layers_count = {}
for rule in master_rules:
    layer = rule.get("Layer", "Unknown")
    layers_count[layer] = layers_count.get(layer, 0) + 1

row_offset += 2
summary_sheet.cell(row=row_offset, column=1).value = "Rules by Layer"
summary_sheet.cell(row=row_offset, column=1).font = Font(bold=True)
row_offset += 1

for layer, count in sorted(layers_count.items()):
    summary_sheet.cell(row=row_offset, column=1).value = layer
    summary_sheet.cell(row=row_offset, column=2).value = count
    row_offset += 1

docs_count = {}
for rule in master_rules:
    doc = rule.get("Source_Document", "Unknown")
    docs_count[doc] = docs_count.get(doc, 0) + 1

row_offset += 2
summary_sheet.cell(row=row_offset, column=1).value = "Rules by Source Document"
summary_sheet.cell(row=row_offset, column=1).font = Font(bold=True)
row_offset += 1

for doc, count in sorted(docs_count.items()):
    summary_sheet.cell(row=row_offset, column=1).value = doc
    summary_sheet.cell(row=row_offset, column=2).value = count
    row_offset += 1

# Add metadata footer
footer_row = row_offset + 3
summary_sheet.cell(row=footer_row, column=1).value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
summary_sheet.cell(row=footer_row+1, column=1).value = "Version: 1.0 (Draft)"
summary_sheet.cell(row=footer_row+2, column=1).value = "Status: Preliminary - Requires full extraction from source documents"

# Save workbook
output_path = r"C:\SIA-Project\SIA-DOCS\02_modules\SEM\01_ord\Drafts\SEM_Master_Rules_Sheet_v1.xlsx"
wb.save(output_path)
print("Master Rules Sheet generated successfully")
print("  Output: " + output_path)
print("  Total Rules: " + str(len(master_rules)))
print("  Summary: {} consolidated rules across 8 source documents".format(len(master_rules)))
print("  Layers: 7 (Foundation, Identity-Roles, Actions-Activities, Validation-Integrity, Gamification, Outputs, Exceptions-Escalations)")
