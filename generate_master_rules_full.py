#!/usr/bin/env python3
"""
SEM Master Rules Sheet Generator - Full Version
Systematically extracts and consolidates rules from Architecture v4.2 + SEM ORD-01 through ORD-07
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from datetime import datetime
import os

# File paths
DOCS = {
    "Architecture v4.2": r"C:\SIA-Project\SIA-DOCS\01_platform_core\architecture\frozen\SIA_Overall_Platform_Architecture_v4.2.md",
    "SEM-ORD-01": r"C:\SIA-Project\SIA-DOCS\02_modules\SEM\01_ord\Drafts\SEM-ORD-01_User_Architecture_and_Role_Framework_v1.md",
    "SEM-ORD-02": r"C:\SIA-Project\SIA-DOCS\02_modules\SEM\01_ord\Drafts\SEM-ORD-02_Campus_Operations_and_Field_Representation_v1.md",
    "SEM-ORD-03": r"C:\SIA-Project\SIA-DOCS\02_modules\SEM\01_ord\Drafts\SEM-ORD-03_Forum_and_Content_Engagement_v1.md",
    "SEM-ORD-04": r"C:\SIA-Project\SIA-DOCS\02_modules\SEM\01_ord\Drafts\SEM-ORD-04_Events_System_v1.md",
    "SEM-ORD-05": r"C:\SIA-Project\SIA-DOCS\02_modules\SEM\01_ord\Drafts\SEM-ORD-05_Gamification_Points_and_Badge_Rules_v1.md",
    "SEM-ORD-06": r"C:\SIA-Project\SIA-DOCS\02_modules\SEM\01_ord\Drafts\SEM-ORD-06_Certification_and_Recognition_v1.md",
    "SEM-ORD-07": r"C:\SIA-Project\SIA-DOCS\02_modules\SEM\01_ord\Drafts\SEM-ORD-07_Platform_Infrastructure_Forms_Dashboards_Onboarding_v1.md",
}

# Color mapping for rule types
COLOR_MAP = {
    "MUST": "FF0000",           # Red
    "NEVER": "8B0000",          # Dark Red
    "RULE": "0070C0",           # Blue
    "POLICY": "00B050",         # Green
    "DEF": "D9D9D9",            # Light Grey
    "NOTE": "E8E8E8",           # Very Light Grey
    "ESCALATION": "FF9900",     # Orange
    "EXCEPTION": "FFC000",      # Yellow
    "MAY": "B4C7E7",            # Light Blue
}

# Layer classification mapping
LAYER_MAPPING = {
    # Foundation layer
    "ARCH": "Foundation",
    "Architecture": "Foundation",
    "Platform": "Foundation",

    # Identity-Roles layer
    "Role": "Identity-Roles",
    "User": "Identity-Roles",
    "Member": "Identity-Roles",
    "Rep": "Identity-Roles",
    "Identity": "Identity-Roles",
    "Scholar": "Identity-Roles",
    "Intern": "Identity-Roles",
    "Alumni": "Identity-Roles",

    # Actions-Activities layer
    "Event": "Actions-Activities",
    "Campus": "Actions-Activities",
    "Activity": "Actions-Activities",
    "Forum": "Actions-Activities",
    "Book": "Actions-Activities",
    "Task": "Actions-Activities",
    "Post": "Actions-Activities",
    "Discussion": "Actions-Activities",

    # Validation-Integrity layer
    "Validation": "Validation-Integrity",
    "Integrity": "Validation-Integrity",
    "Verification": "Validation-Integrity",
    "Approval": "Validation-Integrity",
    "Report": "Validation-Integrity",
    "KYC": "Validation-Integrity",
    "Credit": "Validation-Integrity",
    "Score": "Validation-Integrity",

    # Gamification layer
    "Gamification": "Gamification",
    "Badge": "Gamification",
    "Points": "Gamification",
    "Leaderboard": "Gamification",

    # Outputs layer
    "Dashboard": "Outputs",
    "Certificate": "Outputs",
    "Directory": "Outputs",
    "Form": "Outputs",
    "Onboarding": "Outputs",
    "News": "Outputs",

    # Exceptions-Escalations layer
    "Exception": "Exceptions-Escalations",
    "Escalation": "Exceptions-Escalations",
}

def determine_layer(rule_id, title, description):
    """Determine layer based on rule ID and content"""
    # Try by rule ID pattern
    if rule_id:
        for keyword, layer in LAYER_MAPPING.items():
            if keyword.lower() in rule_id.lower():
                return layer

    # Try by title
    if title:
        for keyword, layer in LAYER_MAPPING.items():
            if keyword.lower() in title.lower():
                return layer

    # Default based on document/section
    return "Actions-Activities"

def get_priority(rule_type):
    """Determine priority based on rule type"""
    if rule_type in ["MUST", "NEVER"]:
        return "CRITICAL"
    elif rule_type in ["RULE", "POLICY"]:
        return "HIGH"
    else:
        return "MEDIUM"

def extract_rules_from_file(doc_name, file_path):
    """Extract all rules from a markdown document"""
    rules = []

    if not os.path.exists(file_path):
        print("Warning: File not found - " + file_path)
        return rules

    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except:
        print("Warning: Could not read file - " + file_path)
        return rules

    # Pattern for rule IDs (e.g., **SEM-ORD-01.1.001** or **ARCH-16**)
    # Capture rule type and text after
    pattern = r'\*\*([A-Z]+-[A-Z0-9.-]+)\*?\*?\s*`([A-Z]+)`\s*\n([^\n]*)\n\n([^*]*?)(?=\n\n##|\n\n\*\*|\Z)'

    matches = re.finditer(pattern, content, re.MULTILINE | re.DOTALL)

    for match in matches:
        rule_id = match.group(1)
        rule_type = match.group(2)
        title = match.group(3).strip()
        description = match.group(4).strip()

        if not title or not description:
            continue

        # Clean up description
        description = description.replace('\n\n', ' ').replace('\n', ' ').strip()
        if len(description) > 300:
            description = description[:297] + "..."

        rule = {
            "Rule_ID": rule_id,
            "Rule_Type": rule_type,
            "Title": title,
            "Description": description,
            "Source": doc_name,
        }

        rules.append(rule)

    return rules

# Extract rules from all documents
all_rules = []
rule_counter = 0

print("Extracting rules from source documents...")
print("-" * 50)

for doc_name, file_path in DOCS.items():
    extracted = extract_rules_from_file(doc_name, file_path)
    print("  {}: {} rules extracted".format(doc_name, len(extracted)))
    all_rules.extend(extracted)

print("-" * 50)
print("Total rules extracted: " + str(len(all_rules)))

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Create Summary Sheet
summary_sheet = wb.create_sheet("Summary", 0)
rules_sheet = wb.create_sheet("Master Rules", 1)

# Define columns
columns = [
    "Master_Rule_ID",
    "Rule_Type",
    "Priority",
    "Layer",
    "Source_Document",
    "Source_Rule_ID",
    "Rule_Title",
    "Rule_Description",
    "Applies_To",
    "Conflict_Flag",
    "Gap_Flag",
    "Edge_Case_Flag",
    "Depends_On",
    "Impacts",
    "Duplicate_Group_ID",
    "Dedup_Status",
    "Implementation_Notes",
    "Cross_Reference",
    "Requires_Config",
    "Scope_Limitation",
    "Authority_Level",
    "Last_Modified",
]

# Set up header row
for col_idx, col_name in enumerate(columns, 1):
    cell = rules_sheet.cell(row=1, column=col_idx)
    cell.value = col_name
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Set column widths
column_widths = [12, 11, 10, 17, 14, 18, 20, 30, 18, 11, 9, 11, 14, 14, 14, 14, 18, 18, 14, 14, 14, 11]
for col_idx, width in enumerate(column_widths, 1):
    rules_sheet.column_dimensions[get_column_letter(col_idx)].width = width

# Populate rules sheet
for row_idx, rule in enumerate(all_rules, 2):
    mr_id = "MR-{:04d}".format(row_idx - 1)
    layer = determine_layer(rule["Rule_ID"], rule["Title"], rule["Description"])
    priority = get_priority(rule["Rule_Type"])

    row_data = {
        "Master_Rule_ID": mr_id,
        "Rule_Type": rule["Rule_Type"],
        "Priority": priority,
        "Layer": layer,
        "Source_Document": rule["Source"],
        "Source_Rule_ID": rule["Rule_ID"],
        "Rule_Title": rule["Title"],
        "Rule_Description": rule["Description"],
        "Applies_To": "",
        "Conflict_Flag": "No",
        "Gap_Flag": "No",
        "Edge_Case_Flag": "No",
        "Depends_On": "",
        "Impacts": "",
        "Duplicate_Group_ID": "None",
        "Dedup_Status": "Master",
        "Implementation_Notes": "",
        "Cross_Reference": rule["Rule_ID"],
        "Requires_Config": "",
        "Scope_Limitation": "v1+",
        "Authority_Level": "Frozen",
        "Last_Modified": "2026-04-13",
    }

    # Write row data
    for col_idx, col_name in enumerate(columns, 1):
        cell = rules_sheet.cell(row=row_idx, column=col_idx)
        cell.value = row_data.get(col_name, "")

        # Apply color based on rule type
        rule_type = rule.get("Rule_Type", "")
        if rule_type in COLOR_MAP:
            cell.fill = PatternFill(start_color=COLOR_MAP[rule_type], end_color=COLOR_MAP[rule_type], fill_type="solid")
            if rule_type in ["MUST", "NEVER"]:
                cell.font = Font(color="FFFFFF", size=9)

        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )

# Freeze header
rules_sheet.freeze_panes = "A2"

# Create Summary Sheet
summary_sheet.column_dimensions["A"].width = 30
summary_sheet.column_dimensions["B"].width = 20

title_cell = summary_sheet.cell(row=1, column=1)
title_cell.value = "SEM Master Rules Sheet - Summary"
title_cell.font = Font(bold=True, size=14, color="1F4E78")

summary_sheet.cell(row=3, column=1).value = "Consolidation Summary"
summary_sheet.cell(row=3, column=1).font = Font(bold=True, size=12)

summary_sheet.cell(row=4, column=1).value = "Total Rules Extracted:"
summary_sheet.cell(row=4, column=2).value = len(all_rules)

# Rules by type
rule_types = {}
for rule in all_rules:
    rt = rule.get("Rule_Type", "Unknown")
    rule_types[rt] = rule_types.get(rt, 0) + 1

summary_sheet.cell(row=6, column=1).value = "Rules by Type"
summary_sheet.cell(row=6, column=1).font = Font(bold=True)

row_offset = 7
for rule_type in sorted(rule_types.keys()):
    count = rule_types[rule_type]
    summary_sheet.cell(row=row_offset, column=1).value = rule_type
    summary_sheet.cell(row=row_offset, column=2).value = count
    if rule_type in COLOR_MAP:
        summary_sheet.cell(row=row_offset, column=1).fill = PatternFill(
            start_color=COLOR_MAP[rule_type], end_color=COLOR_MAP[rule_type], fill_type="solid"
        )
    row_offset += 1

# Rules by source document
row_offset += 2
summary_sheet.cell(row=row_offset, column=1).value = "Rules by Source Document"
summary_sheet.cell(row=row_offset, column=1).font = Font(bold=True)
row_offset += 1

docs_count = {}
for rule in all_rules:
    doc = rule.get("Source", "Unknown")
    docs_count[doc] = docs_count.get(doc, 0) + 1

for doc in sorted(docs_count.keys()):
    summary_sheet.cell(row=row_offset, column=1).value = doc
    summary_sheet.cell(row=row_offset, column=2).value = docs_count[doc]
    row_offset += 1

# Add footer
row_offset += 2
summary_sheet.cell(row=row_offset, column=1).value = "Generated: " + datetime.now().strftime('%Y-%m-%d %H:%M:%S')
summary_sheet.cell(row=row_offset+1, column=1).value = "Version: 1.0 (Draft - Extracted Rules)"
summary_sheet.cell(row=row_offset+2, column=1).value = "Status: Preliminary consolidation - Review for conflicts and gaps"

# Save workbook
output_path = r"C:\SIA-Project\SIA-DOCS\02_modules\SEM\01_ord\Drafts\SEM_Master_Rules_Sheet_v1.xlsx"
wb.save(output_path)

print("")
print("Master Rules Sheet generated successfully")
print("  Output: " + output_path)
print("  Total Rules: " + str(len(all_rules)))
print("  Columns: 22 (comprehensive rule metadata)")
print("  Sheets: Summary + Master Rules")
print("  Status: READY FOR REVIEW")
